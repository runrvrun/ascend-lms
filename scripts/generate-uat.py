"""
Generate UAT script Excel workbook for Ascend LMS.
Run: python3 scripts/generate-uat.py
Output: UAT_Script_Ascend_LMS.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────────
NAVY      = "1E3A6E"
MID_BLUE  = "2563EB"
LIGHT_BLUE= "DBEAFE"
PURPLE    = "7C3AED"
LIGHT_PURPLE="EDE9FE"
GREEN     = "16A34A"
LIGHT_GREEN="DCFCE7"
ORANGE    = "EA580C"
LIGHT_ORANGE="FFEDD5"
RED       = "DC2626"
LIGHT_RED = "FEE2E2"
SLATE     = "475569"
LIGHT_SLATE="F1F5F9"
WHITE     = "FFFFFF"
YELLOW    = "CA8A04"
LIGHT_YELLOW="FEF9C3"
TEAL      = "0F766E"
LIGHT_TEAL="CCFBF1"
GRAY_BORDER="CBD5E1"

def side(color=GRAY_BORDER, style="thin"):
    return Side(border_style=style, color=color)

def border(all=True, thick=False):
    s = side(GRAY_BORDER, "medium" if thick else "thin")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=None, size=10, name="Calibri", italic=False):
    return Font(bold=bold, color=color or "000000", size=size, name=name, italic=italic)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_cell(cell, fg=None, bold=False, color=None, h="left", v="center",
               wrap=True, size=10, border_=True, italic=False):
    if fg:
        cell.fill = fill(fg)
    cell.font = font(bold=bold, color=color or "000000", size=size, italic=italic)
    cell.alignment = align(h=h, v=v, wrap=wrap)
    if border_:
        cell.border = border()

def write(ws, row, col, value, fg=None, bold=False, color=None,
          h="left", v="center", wrap=True, size=10, border_=True, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    style_cell(c, fg=fg, bold=bold, color=color, h=h, v=v,
               wrap=wrap, size=size, border_=border_, italic=italic)
    return c

# ── Test case data ──────────────────────────────────────────────────────────────
# Each entry: (tc_id, phase, role, feature, test_step, expected_result, priority)
TEST_CASES = [
    # ─── PHASE 1: SETUP ───────────────────────────────────────────────────────
    ("TC-001", "1 – Setup", "Admin", "User Account Activation",
     "Receive activation email → Click activation link → Set password (min. 8 chars) → Click Activate account",
     "Password is set, page shows 'Password set successfully!' and redirects to sign-in page.",
     "Critical"),

    ("TC-002", "1 – Setup", "Admin", "Sign In (Microsoft SSO)",
     "Navigate to the app → Click 'Sign in with Microsoft' → Select existing Microsoft account from picker",
     "User is signed in and redirected to the Home dashboard without re-entering credentials.",
     "Critical"),

    ("TC-003", "1 – Setup", "Admin", "Sign In (Password)",
     "Navigate to sign-in page → Enter email and password → Click Sign in",
     "User is signed in and redirected to the Home dashboard.",
     "Critical"),

    # ─── PHASE 2: ADMIN – CONTENT SETUP ───────────────────────────────────────
    ("TC-004", "2 – Admin Setup", "Admin", "Create Office",
     "Admin → Offices → Add Office → Enter office name → Save",
     "Office appears in the office list.",
     "High"),

    ("TC-005", "2 – Admin Setup", "Admin", "Create Topic",
     "Admin → Topics → Add Topic → Enter topic name → Save",
     "Topic appears in the topic list.",
     "High"),

    ("TC-006", "2 – Admin Setup", "Admin", "Create User (manual)",
     "Admin → Users → Add User → Fill name, email, division, title, office → Save",
     "User appears in user list with status Active.",
     "Critical"),

    ("TC-007", "2 – Admin Setup", "Admin", "Bulk Import Users",
     "Admin → Users → Import → Download template → Fill in at least 3 rows (include Manager Email column) → Upload file",
     "Users are created; any assigned managers are linked. Success count shown.",
     "High"),

    ("TC-008", "2 – Admin Setup", "Admin", "Send Activation Email",
     "Admin → Users → Actions (user row) → Send Activation Email",
     "Confirmation shown. User receives activation email within a few minutes.",
     "Critical"),

    ("TC-009", "2 – Admin Setup", "Admin", "Assign Roles to User",
     "Admin → Users → Actions → Assign Roles → Tick 'Manager' and/or 'Trainer' → Save",
     "Role badge appears on the user row. User can now access that role's section.",
     "Critical"),

    ("TC-010", "2 – Admin Setup", "Admin", "Assign Manager to User",
     "Admin → Users → Actions → Assign Managers → Type manager name → Click to add → Close modal",
     "Manager chip appears. User is now linked to that manager.",
     "High"),

    ("TC-011", "2 – Admin Setup", "Admin", "Assign User to Cohort",
     "Admin → Users → Actions → Assign Cohorts → Select cohort → Save",
     "User is listed as a member of the cohort.",
     "Medium"),

    ("TC-012", "2 – Admin Setup", "Admin", "Create Cohort",
     "Admin → Cohorts → Add Cohort → Enter cohort name → Save",
     "Cohort appears in cohort list.",
     "Medium"),

    ("TC-013", "2 – Admin Setup", "Admin", "Create Course (no test/assignment)",
     "Admin → Courses → Add Course → Enter name, description, topic → Save",
     "Course is created with status Published. Appears in course list.",
     "Critical"),

    ("TC-014", "2 – Admin Setup", "Admin", "Assign Trainer to Course",
     "Admin → Courses → [Course] → Manage Contents → Trainers section → Add Trainer → Search trainer name → Assign",
     "Trainer appears in the trainers list for that course.",
     "High"),

    ("TC-015", "2 – Admin Setup", "Admin", "Add Text Content to Course",
     "Admin → Courses → [Course] → Manage Contents → Content tab → Add Content → Type: Text → Enter title and body → Save",
     "Text content item appears in the course content list in correct order.",
     "Critical"),

    ("TC-016", "2 – Admin Setup", "Admin", "Add Video Content to Course",
     "Admin → Courses → [Course] → Content tab → Add Content → Type: Video → Enter title, URL, duration → Save",
     "Video content item appears with correct duration.",
     "High"),

    ("TC-017", "2 – Admin Setup", "Admin", "Add Link Content to Course",
     "Admin → Courses → [Course] → Content tab → Add Content → Type: Link → Enter title and URL → Save",
     "Link content item appears in course.",
     "Medium"),

    ("TC-018", "2 – Admin Setup", "Admin", "Create Test for Course",
     "Admin → Courses → [Course] → Test tab → Create Test → Set pass threshold (e.g. 70%) → Add at least 3 questions (mix types) → Save",
     "Test is saved with the correct threshold and all questions visible.",
     "Critical"),

    ("TC-019", "2 – Admin Setup", "Admin", "Create Assignment for Course",
     "Admin → Courses → [Course] → Assignment tab → Create Assignment → Enter description and SharePoint URL → Save",
     "Assignment is saved and visible with the submission URL.",
     "High"),

    ("TC-020", "2 – Admin Setup", "Admin", "Create Pathway",
     "Admin → Pathways → Add Pathway → Enter name, description, tags → Save",
     "Pathway is created with status Published.",
     "Critical"),

    ("TC-021", "2 – Admin Setup", "Admin", "Add Courses to Pathway",
     "Admin → Pathways → [Pathway] → Add Course → Select course → Set order and points → Save",
     "Course appears in the pathway's course list with correct order and points.",
     "Critical"),

    ("TC-022", "2 – Admin Setup", "Admin", "Set Pathway to Require Approval",
     "Admin → Pathways → [Pathway] → Enable 'Requires Approval' toggle → Save",
     "Pathway shows 'Requires Approval' badge. Learners must request enrollment.",
     "High"),

    ("TC-023", "2 – Admin Setup", "Admin", "Assign Pathway to Cohort",
     "Admin → Cohorts → [Cohort] → Pathways tab → Assign Pathway → Select pathway",
     "Pathway is linked to cohort. All cohort members are auto-enrolled.",
     "High"),

    ("TC-024", "2 – Admin Setup", "Admin", "Draft / Publish Course Toggle",
     "Admin → Courses → [Course] → Manage Contents → Click status toggle to set to Draft → Verify learners cannot see it → Toggle back to Published",
     "Drafted course is hidden from learners. Published course is visible.",
     "Medium"),

    # ─── PHASE 3: SME WORKFLOW ─────────────────────────────────────────────────
    ("TC-025", "3 – SME", "SME", "Assign SME to Topic",
     "Admin → Topics → [Topic] → SME field → Search user with SME role → Assign",
     "SME is listed under that topic.",
     "High"),

    ("TC-026", "3 – SME", "SME", "View Assigned Topics",
     "Sign in as SME → Subject Matter Expert → My Topics",
     "SME sees all topics they are assigned to and the courses within each topic.",
     "High"),

    ("TC-027", "3 – SME", "SME", "Create Course as SME",
     "SME → My Topics → [Topic] → New Course → Fill details → Save",
     "Course is created under that topic and appears in the course list.",
     "High"),

    ("TC-028", "3 – SME", "SME", "Manage Content as SME",
     "SME → My Topics → [Course] → Add content item → Save",
     "Content item appears in course. Same capability as Trainer.",
     "Medium"),

    # ─── PHASE 4: TRAINER WORKFLOW ─────────────────────────────────────────────
    ("TC-029", "4 – Trainer", "Trainer", "View Assigned Courses",
     "Sign in as Trainer → Trainer → My Courses",
     "Trainer sees only courses they are assigned to.",
     "High"),

    ("TC-030", "4 – Trainer", "Trainer", "Edit Course Content",
     "Trainer → My Courses → [Course] → Content tab → Edit an existing content item → Save",
     "Content is updated with the new values.",
     "High"),

    ("TC-031", "4 – Trainer", "Trainer", "Edit Test Questions",
     "Trainer → My Courses → [Course] → Test tab → Edit a question → Save",
     "Question is updated correctly.",
     "High"),

    ("TC-032", "4 – Trainer", "Trainer", "Grade Assignment Submission",
     "Trainer → My Courses → [Course] → Submissions tab → Click Grade on a submitted assignment → Select Passed → Add note → Save",
     "Submission status changes to Passed. Learner receives a notification.",
     "Critical"),

    ("TC-033", "4 – Trainer", "Trainer", "Grade Assignment as Failed",
     "Trainer → My Courses → [Course] → Submissions tab → Click Grade → Select Failed → Add note → Save",
     "Submission status changes to Failed. Learner receives a notification.",
     "High"),

    ("TC-034", "4 – Trainer", "Trainer", "View Course Feedback",
     "Trainer → My Courses → [Course] → Feedback tab",
     "Average star rating and individual learner comments are shown.",
     "Medium"),

    # ─── PHASE 5: LEARNER SELF-ENROLL FLOW ────────────────────────────────────
    ("TC-035", "5 – Learner (Self-Enroll)", "Learner", "View Home Dashboard",
     "Sign in as Learner → Home page",
     "Dashboard shows enrolled pathways with progress bars, points, streak, and leaderboard rank.",
     "Critical"),

    ("TC-036", "5 – Learner (Self-Enroll)", "Learner", "Browse Pathways",
     "Learner → Pathways → Browse published pathways",
     "All published pathways are shown with name, description, and tags.",
     "Critical"),

    ("TC-037", "5 – Learner (Self-Enroll)", "Learner", "Self-Enroll in Pathway (no approval)",
     "Learner → Pathways → [Pathway without approval] → Enroll",
     "Learner is immediately enrolled. Pathway appears on Home dashboard.",
     "Critical"),

    ("TC-038", "5 – Learner (Self-Enroll)", "Learner", "Request Enrollment in Pathway (approval required)",
     "Learner → Pathways → [Pathway with approval] → Request Enrollment → Add note → Submit",
     "Request is submitted. Learner sees 'Pending' status. Manager receives notification.",
     "Critical"),

    ("TC-039", "5 – Learner (Self-Enroll)", "Learner", "Open Course Content (Text)",
     "Learner → [Enrolled Pathway] → Open course → Click text content item",
     "Text content is displayed in the main panel.",
     "Critical"),

    ("TC-040", "5 – Learner (Self-Enroll)", "Learner", "Watch Video Content",
     "Learner → Course → Click video content → Watch at least 75% of the video",
     "Progress bar advances. 'Mark as Completed' button becomes available after 75%.",
     "Critical"),

    ("TC-041", "5 – Learner (Self-Enroll)", "Learner", "Mark Content as Completed",
     "Learner → Course → After watching/reading content → Click 'Mark as Completed'",
     "Content item shows green checkmark. Progress counter updates.",
     "Critical"),

    ("TC-042", "5 – Learner (Self-Enroll)", "Learner", "Take and Pass Test",
     "Learner completes all content → Test section appears → Start Test → Answer all questions correctly → Submit",
     "Score shown. Status: Passed. Course marked as complete. Points awarded.",
     "Critical"),

    ("TC-043", "5 – Learner (Self-Enroll)", "Learner", "Take and Fail Test, Then Retake",
     "Learner → Start Test → Answer incorrectly below threshold → Submit → Retake test → Pass",
     "First attempt shows Failed. Retake is available. Second attempt shows Passed.",
     "High"),

    ("TC-044", "5 – Learner (Self-Enroll)", "Learner", "Submit Assignment",
     "Learner → Course → Assignment section → Upload file to SharePoint link → Paste submission URL → Submit",
     "Submission is recorded with status 'Submitted'. Trainer receives notification.",
     "Critical"),

    ("TC-045", "5 – Learner (Self-Enroll)", "Learner", "Leave Course Feedback",
     "Learner → [Completed Course with feedback enabled] → Feedback section → Select star rating → Add comment → Submit",
     "Feedback is saved and shown in the Trainer's Feedback tab.",
     "Medium"),

    ("TC-046", "5 – Learner (Self-Enroll)", "Learner", "Comment on Content",
     "Learner → Course content → Discussion section → Type comment → Post",
     "Comment appears in the discussion thread.",
     "Medium"),

    ("TC-047", "5 – Learner (Self-Enroll)", "Learner", "Reply to Comment",
     "Learner → Discussion → Click Reply on an existing comment → Type reply → Post",
     "Reply appears nested under the parent comment. Author is notified.",
     "Medium"),

    ("TC-048", "5 – Learner (Self-Enroll)", "Learner", "View Learning History",
     "Learner → sidebar → Learning History",
     "Completed courses grouped by pathway are shown, with completion date and test score.",
     "High"),

    ("TC-049", "5 – Learner (Self-Enroll)", "Learner", "Download Certificate",
     "Learner completes ALL courses in a pathway → Open pathway → Click 'Get Certificate' → Certificate page → Click 'Download Certificate'",
     "Certificate page shows learner name, pathway name, and completion date. Browser print dialog opens for PDF download.",
     "Critical"),

    # ─── PHASE 6: MANAGER-APPROVED ENROLLMENT ─────────────────────────────────
    ("TC-050", "6 – Manager Approval", "Manager", "View Team Members",
     "Sign in as Manager → Manager → Professionals",
     "All users assigned to this manager are listed with name, division, pathways count, and points.",
     "Critical"),

    ("TC-051", "6 – Manager Approval", "Manager", "Approve Pathway Request",
     "Manager → Pathway Requests → Find pending request → Click Approve",
     "Request status changes to Approved. Learner receives approval notification and is enrolled.",
     "Critical"),

    ("TC-052", "6 – Manager Approval", "Manager", "Reject Pathway Request",
     "Manager → Pathway Requests → Find pending request → Click Reject → Enter rejection reason → Confirm",
     "Request status changes to Rejected. Learner receives rejection notification with reason.",
     "High"),

    ("TC-053", "6 – Manager Approval", "Manager", "Assign Pathway to Team Member",
     "Manager → Professionals → [Team Member] → Learning tab → Assign Pathway → Select pathway → Set deadline → Confirm",
     "Team member is enrolled. Pathway appears in their Home dashboard. Deadline is visible.",
     "High"),

    ("TC-054", "6 – Manager Approval", "Manager", "Update Enrollment Deadline",
     "Manager → Professionals → [Team Member] → Learning tab → Click edit icon next to enrollment → Update deadline → Save",
     "Deadline is updated for that enrollment.",
     "Medium"),

    ("TC-055", "6 – Manager Approval", "Manager", "View Team Member Progress",
     "Manager → Professionals → [Team Member] → View Progress",
     "Detailed view shows enrolled pathways, course completion per pathway, and overall progress.",
     "High"),

    ("TC-056", "6 – Manager Approval", "Manager", "Confirm Growth Plan Goal",
     "Manager receives notification of completed goal → Go to team member's Growth Plan tab → Click Confirm on a completed goal",
     "Goal is marked as Confirmed. Team member receives notification.",
     "High"),

    ("TC-057", "6 – Manager Approval", "Manager", "View Team Analytics",
     "Manager → Analytics → Review metrics and use filters",
     "Shows team size, enrolments, completion rate, average points, and per-division breakdown.",
     "Medium"),

    # ─── PHASE 7: GROWTH PLAN ─────────────────────────────────────────────────
    ("TC-058", "7 – Growth Plan", "Learner", "Create Growth Plan Goal",
     "Learner → My Growth Plan → Add Item → Enter goal title → Optionally link a pathway → Save",
     "Goal appears in the growth plan list with status In Progress.",
     "High"),

    ("TC-059", "7 – Growth Plan", "Learner", "Mark Goal as Completed",
     "Learner → My Growth Plan → Click checkmark on an in-progress goal",
     "Goal status changes to Completed. Manager receives notification.",
     "High"),

    ("TC-060", "7 – Growth Plan", "Learner", "View Confirmed Goals",
     "Learner → My Growth Plan → Check confirmed items",
     "Confirmed goals show green 'Confirmed' badge with confirmation date.",
     "Medium"),

    # ─── PHASE 8: NOTIFICATIONS ────────────────────────────────────────────────
    ("TC-061", "8 – Notifications", "Learner", "Receive Enrollment Approval Notification",
     "After manager approves enrollment → Learner checks notifications (bell icon)",
     "Unread count badge appears. Notification reads: pathway approved. Clicking navigates to the pathway.",
     "High"),

    ("TC-062", "8 – Notifications", "Learner", "Receive Assignment Graded Notification",
     "After trainer grades assignment → Learner checks notifications",
     "Notification shows grade result (Passed/Failed). Clicking navigates to the course.",
     "High"),

    ("TC-063", "8 – Notifications", "Manager", "Receive Pathway Request Notification",
     "After learner requests enrollment → Manager checks notifications",
     "Notification shows learner name and pathway. Clicking navigates to Pathway Requests page.",
     "High"),

    ("TC-064", "8 – Notifications", "Learner", "Mark Notifications as Read",
     "Learner → Notifications page → View all notifications",
     "Unread notifications are highlighted. After visiting the page, unread count resets.",
     "Medium"),

    # ─── PHASE 9: ANALYTICS ────────────────────────────────────────────────────
    ("TC-065", "9 – Analytics", "Admin", "View Admin Analytics",
     "Admin → Analytics → Review default view",
     "Shows total users, enrolments, completion rate, avg points, top pathways, and global leaderboard.",
     "High"),

    ("TC-066", "9 – Analytics", "Admin", "Filter Analytics by Division",
     "Admin → Analytics → Select a division from the filter",
     "All metrics update to reflect only users in that division.",
     "Medium"),

    ("TC-067", "9 – Analytics", "Admin", "View Leaderboard",
     "Learner → Home dashboard → Leaderboard widget",
     "Top users ranked by total points are displayed. Current user's rank is highlighted.",
     "Medium"),

    # ─── PHASE 10: SETTINGS & MISC ────────────────────────────────────────────
    ("TC-068", "10 – Settings", "Any", "Change Password",
     "Sidebar → Settings gear → Change Password → Enter current password → Enter new password → Confirm → Save",
     "Password is updated. Next sign-in requires the new password.",
     "High"),

    ("TC-069", "10 – Settings", "Any", "View User Guide",
     "Sidebar → Help → Select a role tab → Expand an accordion section",
     "Step-by-step instructions are shown for the selected role and section.",
     "Low"),

    ("TC-070", "10 – Settings", "Admin", "Edit User Details",
     "Admin → Users → Actions → Edit → Update name or division → Save",
     "User details are updated in the list.",
     "Medium"),

    ("TC-071", "10 – Settings", "Admin", "Delete User",
     "Admin → Users → Actions → Delete → Confirm",
     "User is removed from the list.",
     "Medium"),
]

# ── Priority fill colors ────────────────────────────────────────────────────────
PRIORITY_STYLE = {
    "Critical": (LIGHT_RED,   RED,    "Critical"),
    "High":     (LIGHT_ORANGE, ORANGE, "High"),
    "Medium":   (LIGHT_YELLOW, YELLOW, "Medium"),
    "Low":      (LIGHT_GREEN,  GREEN,  "Low"),
}

# ── Phase header colors ─────────────────────────────────────────────────────────
PHASE_COLORS = {
    "1 – Setup":                     (NAVY,    WHITE),
    "2 – Admin Setup":               ("1D4ED8", WHITE),
    "3 – SME":                       (PURPLE,   WHITE),
    "4 – Trainer":                   (TEAL,     WHITE),
    "5 – Learner (Self-Enroll)":     (GREEN,    WHITE),
    "6 – Manager Approval":          ("7C3AED", WHITE),
    "7 – Growth Plan":               ("0891B2", WHITE),
    "8 – Notifications":             ("B45309", WHITE),
    "9 – Analytics":                 (SLATE,    WHITE),
    "10 – Settings":                 ("374151", WHITE),
}

def build_cover(wb):
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 3

    # Banner
    for r in range(1, 12):
        for c in range(1, 5):
            ws.cell(r, c).fill = fill(NAVY)

    ws.merge_cells("B2:C3")
    c = ws["B2"]
    c.value = "ASCEND LMS"
    c.font = Font(bold=True, size=28, color=WHITE, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B4:C5")
    c = ws["B4"]
    c.value = "User Acceptance Testing (UAT) Script"
    c.font = Font(bold=False, size=16, color="BFDBFE", name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B7:C7")
    c = ws["B7"]
    c.value = "Full-Flow Test: Course Creation → Pathway Enrollment → Certificate Download"
    c.font = Font(bold=False, size=11, color="93C5FD", name="Calibri", italic=True)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Metadata table
    meta = [
        ("Project",    "Ascend LMS"),
        ("Version",    "1.0"),
        ("Prepared by",""),
        ("Date",       ""),
        ("Environment",""),
        ("Tested by",  ""),
    ]
    row = 13
    for label, val in meta:
        ws.row_dimensions[row].height = 22
        c = ws.cell(row, 2, label)
        c.font = Font(bold=True, size=10, name="Calibri", color=NAVY)
        c.fill = fill(LIGHT_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(bottom=side(GRAY_BORDER))
        c2 = ws.cell(row, 3, val)
        c2.font = Font(size=10, name="Calibri")
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c2.border = Border(bottom=side(GRAY_BORDER))
        row += 1

    # Legend
    row += 2
    ws.merge_cells(f"B{row}:C{row}")
    c = ws.cell(row, 2, "Result Legend")
    c.font = Font(bold=True, size=11, color=NAVY, name="Calibri")
    row += 1

    legend = [
        ("PASS",    LIGHT_GREEN,  GREEN),
        ("FAIL",    LIGHT_RED,    RED),
        ("BLOCK",   LIGHT_ORANGE, ORANGE),
        ("N/A",     LIGHT_SLATE,  SLATE),
    ]
    for label, bg, fg in legend:
        ws.row_dimensions[row].height = 20
        c = ws.cell(row, 2, label)
        c.fill = fill(bg)
        c.font = Font(bold=True, size=10, color=fg, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border()
        row += 1

    # Instructions
    row += 2
    ws.merge_cells(f"B{row}:C{row}")
    c = ws.cell(row, 2, "How to use this script")
    c.font = Font(bold=True, size=11, color=NAVY, name="Calibri")
    row += 1

    instructions = [
        "1. Work through each test case in tab order.",
        "2. Execute the Test Steps exactly as written.",
        "3. Compare the actual outcome with the Expected Result.",
        "4. Record PASS / FAIL / BLOCK / N/A in the 'Result' column.",
        "5. Add notes in the 'Actual Result / Notes' column for any FAIL or BLOCK.",
        "6. Use the 'Tester' and 'Date Tested' columns to track who ran each test.",
    ]
    for inst in instructions:
        ws.row_dimensions[row].height = 20
        c = ws.cell(row, 2, inst)
        c.font = Font(size=10, name="Calibri", color=SLATE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        row += 1


def build_test_sheet(wb):
    ws = wb.create_sheet("Test Cases")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Column widths
    col_widths = [6, 8, 22, 16, 22, 38, 40, 12, 12, 20, 28]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Header rows ──────────────────────────────────────────────────────────
    HEADERS = [
        "#", "ID", "Phase", "Role", "Feature",
        "Test Steps", "Expected Result",
        "Priority", "Result", "Tester / Date", "Actual Result / Notes"
    ]

    # Row 1: big title bar
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "ASCEND LMS – UAT Test Cases"
    c.fill = fill(NAVY)
    c.font = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Row 2: column headers
    ws.row_dimensions[2].height = 28
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(2, col, h)
        c.fill = fill(MID_BLUE)
        c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border(thick=True)

    # ── Data rows ────────────────────────────────────────────────────────────
    current_phase = None
    data_row = 3
    seq = 0

    for tc in TEST_CASES:
        tc_id, phase, role, feature, steps, expected, priority = tc
        seq += 1

        # Phase header row
        if phase != current_phase:
            current_phase = phase
            ws.merge_cells(f"A{data_row}:K{data_row}")
            bg, fg_c = PHASE_COLORS.get(phase, (SLATE, WHITE))
            c = ws.cell(data_row, 1, f"  ▸  {phase}")
            c.fill = fill(bg)
            c.font = Font(bold=True, size=10, color=fg_c, name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[data_row].height = 22
            data_row += 1

        # Test case row
        ws.row_dimensions[data_row].height = 60
        row_bg = WHITE if seq % 2 == 0 else LIGHT_SLATE

        values = [seq, tc_id, phase, role, feature, steps, expected, priority, "", "", ""]
        for col, val in enumerate(values, 1):
            c = ws.cell(data_row, col, val)
            c.fill = fill(row_bg)
            c.font = font(size=9)
            c.alignment = align(h="left" if col > 2 else "center", wrap=True)
            c.border = border()

        # Priority cell styling
        pri_bg, pri_fg, _ = PRIORITY_STYLE[priority]
        pc = ws.cell(data_row, 8)
        pc.fill = fill(pri_bg)
        pc.font = Font(bold=True, size=9, color=pri_fg, name="Calibri")
        pc.alignment = Alignment(horizontal="center", vertical="center")

        # Result cell – dropdown hint
        rc = ws.cell(data_row, 9, "")
        rc.fill = fill("FAFAFA")
        rc.alignment = Alignment(horizontal="center", vertical="center")

        data_row += 1

    # Data validation for Result column (col 9)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"PASS,FAIL,BLOCK,N/A"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = f"I3:I{data_row}"
    ws.add_data_validation(dv)


def build_summary(wb):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 3

    # Title
    ws.merge_cells("B1:D1")
    c = ws["B1"]
    c.value = "UAT Summary"
    c.fill = fill(NAVY)
    c.font = Font(bold=True, size=14, color=WHITE, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = [("Phase", MID_BLUE), ("Total TCs", MID_BLUE), ("Priority Breakdown", MID_BLUE)]
    row = 3
    for col, (h, bg) in enumerate(headers, 2):
        c = ws.cell(row, col, h)
        c.fill = fill(bg)
        c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border()
    ws.row_dimensions[row].height = 22
    row += 1

    # Count by phase
    from collections import Counter
    phase_counts = Counter(tc[1] for tc in TEST_CASES)
    phase_priority = {}
    for tc in TEST_CASES:
        p, pri = tc[1], tc[6]
        if p not in phase_priority:
            phase_priority[p] = Counter()
        phase_priority[p][pri] += 1

    for phase in sorted(phase_counts.keys(), key=lambda x: int(x.split()[0])):
        bg, _ = PHASE_COLORS.get(phase, (SLATE, WHITE))
        ws.row_dimensions[row].height = 20
        c = ws.cell(row, 2, phase)
        c.fill = fill(bg)
        c.font = Font(bold=True, size=9, color=WHITE, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border()

        c2 = ws.cell(row, 3, phase_counts[phase])
        c2.font = Font(bold=True, size=10, name="Calibri")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = border()

        ppc = phase_priority.get(phase, Counter())
        pri_str = " | ".join(f"{k}:{v}" for k, v in ppc.items())
        c3 = ws.cell(row, 4, pri_str)
        c3.font = Font(size=9, name="Calibri", color=SLATE)
        c3.alignment = Alignment(horizontal="center", vertical="center")
        c3.border = border()
        row += 1

    # Totals
    ws.row_dimensions[row].height = 22
    c = ws.cell(row, 2, "TOTAL")
    c.fill = fill(NAVY)
    c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border()
    c2 = ws.cell(row, 3, len(TEST_CASES))
    c2.fill = fill(NAVY)
    c2.font = Font(bold=True, size=11, color=WHITE, name="Calibri")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border = border()
    c3 = ws.cell(row, 4, "")
    c3.fill = fill(NAVY)
    c3.border = border()
    row += 2

    # Priority totals
    all_pri = Counter(tc[6] for tc in TEST_CASES)
    c = ws.cell(row, 2, "Priority Summary")
    c.font = Font(bold=True, size=10, color=NAVY, name="Calibri")
    row += 1
    for pri in ["Critical", "High", "Medium", "Low"]:
        bg, fg_c, _ = PRIORITY_STYLE[pri]
        ws.row_dimensions[row].height = 20
        c = ws.cell(row, 2, pri)
        c.fill = fill(bg)
        c.font = Font(bold=True, size=9, color=fg_c, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border()
        c2 = ws.cell(row, 3, all_pri.get(pri, 0))
        c2.font = Font(bold=True, size=10, name="Calibri")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = border()
        row += 1

    # Sign-off table
    row += 2
    ws.merge_cells(f"B{row}:D{row}")
    c = ws.cell(row, 2, "Sign-Off")
    c.font = Font(bold=True, size=11, color=NAVY, name="Calibri")
    row += 1

    signoff_cols = ["Role", "Name", "Signature / Date"]
    for col, h in enumerate(signoff_cols, 2):
        c = ws.cell(row, col, h)
        c.fill = fill(MID_BLUE)
        c.font = Font(bold=True, size=9, color=WHITE, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border()
    row += 1

    signoff_rows = ["UAT Coordinator", "HR Admin", "IT / Developer", "Approver"]
    for r in signoff_rows:
        ws.row_dimensions[row].height = 24
        c = ws.cell(row, 2, r)
        c.font = Font(size=9, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border()
        for col in [3, 4]:
            cc = ws.cell(row, col, "")
            cc.border = border()
        row += 1


# ── Main ────────────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

build_cover(wb)
build_test_sheet(wb)
build_summary(wb)

out_path = "UAT_Script_Ascend_LMS.xlsx"
wb.save(out_path)
print(f"✓ Saved: {out_path}  ({len(TEST_CASES)} test cases)")
